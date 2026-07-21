"""
DNX3 Wave Tracker — Excel → data.js Import Script
===================================================
Reads the daily sequencing Excel file and generates data.js
for the Wave Tracker app.

EXPECTED EXCEL FORMAT
─────────────────────
One sheet named "Sequencing" (or the first sheet if not found).
Each row = one route. Required columns (header names flexible, see COLUMN_MAP):

  | Wave Time  | Route Code | Staging      | DSP  | Staging Group |
  |------------|------------|--------------|------|---------------|
  | 10:50 AM   | CA_A145    | STG-D1-21.6  | DERD | B             |
  | 10:50 AM   | CA_A153    | STG-C1-28.11 | DERD | C             |
  | 11:10 AM   | CA_A104    | STG-D1-21.6  | ALGK | D             |
  ...

STAGING GROUP COLUMN
────────────────────
  B or D  →  green card (Staging B & D)
  A or C  →  red card   (Staging A & C)

  Alternatively, leave this column out and the script will auto-detect
  from the staging code itself (STG-A, STG-B, STG-C, STG-D).

USAGE
─────
  python import_data.py                          # uses sequencing.xlsx in same folder
  python import_data.py myfile.xlsx              # custom file
  python import_data.py myfile.xlsx Sheet2       # custom file + sheet name

OUTPUT
──────
  Overwrites data.js in the same folder as this script.
  A backup of the previous data.js is saved as data.js.bak

DEPENDENCIES
────────────
  pip install openpyxl
"""

import sys
import os
import re
import json
import shutil
from datetime import datetime

# ── Try importing openpyxl ──────────────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed.")
    print("Run:  pip install openpyxl")
    sys.exit(1)

# ── Config ──────────────────────────────────────────────────────────────────

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Default input file (place in same folder as this script)
DEFAULT_EXCEL = os.path.join(SCRIPT_DIR, "sequencing.xlsx")

# Output file
OUTPUT_JS = os.path.join(SCRIPT_DIR, "data.js")

# Column name mapping — add aliases if your Excel uses different headers
# Keys are canonical names, values are lists of accepted header variations (lowercase)
COLUMN_MAP = {
    "wave":    ["wave time", "wave", "departure", "wave_time", "departuretime"],
    "route":   ["route code", "route", "route_code", "routecode", "tour", "tour code"],
    "staging": ["staging", "staging location", "staging_location", "stage", "position"],
    "dsp":     ["dsp", "dsp code", "dsp_code", "provider", "carrier"],
    "group":   ["staging group", "group", "staging_group", "color", "zone"],   # optional
}

# Staging group auto-detection from staging code (used when no group column)
# STG-B or STG-D → green, STG-A or STG-C → red
def detect_group(staging_code: str) -> str:
    """Return 'green' or 'red' based on staging code. Falls back to 'red'."""
    code = str(staging_code).upper()
    match = re.search(r'STG-([ABCD])', code)
    if match:
        letter = match.group(1)
        return "green" if letter in ("B", "D") else "red"
    return "red"

def group_to_color(group_val: str) -> str:
    """Normalise a group column value to 'green' or 'red'."""
    v = str(group_val).strip().upper()
    if v in ("B", "D", "GREEN", "STG-B", "STG-D"):
        return "green"
    if v in ("A", "C", "RED", "STG-A", "STG-C"):
        return "red"
    # Fall back to auto-detect — caller should pass staging code instead
    return None

# ── Time normalisation ───────────────────────────────────────────────────────

def normalise_time(raw) -> str:
    """
    Accept many time formats and return "HH:MM AM/PM".
    Examples: "10:50 AM", "10:50", datetime.time(10,50), 0.451388 (Excel fraction)
    """
    if raw is None:
        return None

    # datetime.time object (openpyxl sometimes returns these)
    if hasattr(raw, 'hour'):
        h, m = raw.hour, raw.minute
        ap = "AM" if h < 12 else "PM"
        h12 = h % 12 or 12
        return f"{h12}:{m:02d} {ap}"

    # Excel stores times as float fractions of a day
    if isinstance(raw, float):
        total_minutes = round(raw * 24 * 60)
        h, m = divmod(total_minutes, 60)
        ap = "AM" if h < 12 else "PM"
        h12 = h % 12 or 12
        return f"{h12}:{m:02d} {ap}"

    s = str(raw).strip()

    # Already formatted "10:50 AM"
    if re.match(r'^\d{1,2}:\d{2}\s*(AM|PM)$', s, re.IGNORECASE):
        parts = s.split()
        return f"{parts[0]} {parts[1].upper()}"

    # 24h "10:50" or "14:30"
    m24 = re.match(r'^(\d{1,2}):(\d{2})$', s)
    if m24:
        h, m = int(m24.group(1)), int(m24.group(2))
        ap = "AM" if h < 12 else "PM"
        h12 = h % 12 or 12
        return f"{h12}:{m:02d} {ap}"

    return s  # return as-is if unrecognised

# ── Column detection ─────────────────────────────────────────────────────────

def find_columns(headers: list) -> dict:
    """Map canonical column names to column indices."""
    headers_lower = [str(h).strip().lower() if h else "" for h in headers]
    result = {}
    for canonical, aliases in COLUMN_MAP.items():
        for i, h in enumerate(headers_lower):
            if h in aliases:
                result[canonical] = i
                break
    return result

# ── Main conversion ──────────────────────────────────────────────────────────

def excel_to_waves(excel_path: str, sheet_name: str = None) -> list:
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # Sheet selection
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    elif "Sequencing" in wb.sheetnames:
        ws = wb["Sequencing"]
    else:
        ws = wb.active
        print(f"  Sheet not specified or not found — using active sheet: '{ws.title}'")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Sheet is empty.")

    # Detect header row (first non-empty row)
    header_row = None
    data_start = 0
    for i, row in enumerate(rows):
        if any(cell is not None for cell in row):
            header_row = row
            data_start = i + 1
            break

    if header_row is None:
        raise ValueError("No header row found.")

    col = find_columns(list(header_row))
    print(f"  Columns detected: {col}")

    required = ["wave", "route", "staging", "dsp"]
    missing_cols = [c for c in required if c not in col]
    if missing_cols:
        raise ValueError(
            f"Required columns not found: {missing_cols}\n"
            f"  Headers seen: {[str(h).strip() for h in header_row if h]}\n"
            f"  Expected one of: { {k: v for k, v in COLUMN_MAP.items() if k in missing_cols} }"
        )

    has_group_col = "group" in col

    # Build wave dict: { "10:50 AM": { "green": [...], "red": [...] } }
    waves_dict = {}
    wave_order = []
    skipped = 0

    for row_num, row in enumerate(rows[data_start:], start=data_start + 2):
        try:
            wave_raw  = row[col["wave"]]
            route_raw = row[col["route"]]
            stg_raw   = row[col["staging"]]
            dsp_raw   = row[col["dsp"]]
        except IndexError:
            skipped += 1
            continue

        # Skip empty rows
        if not any([wave_raw, route_raw, stg_raw, dsp_raw]):
            continue

        wave_time = normalise_time(wave_raw)
        route     = str(route_raw).strip() if route_raw else None
        staging   = str(stg_raw).strip() if stg_raw else None
        dsp       = str(dsp_raw).strip().upper() if dsp_raw else None

        if not all([wave_time, route, staging, dsp]):
            print(f"  ⚠ Row {row_num}: incomplete data — skipping. ({wave_time!r}, {route!r}, {staging!r}, {dsp!r})")
            skipped += 1
            continue

        # Determine green/red
        if has_group_col:
            grp_raw = row[col["group"]]
            color = group_to_color(str(grp_raw)) if grp_raw else None
            if color is None:
                color = detect_group(staging)
        else:
            color = detect_group(staging)

        # Register wave
        if wave_time not in waves_dict:
            waves_dict[wave_time] = {"green": [], "red": []}
            wave_order.append(wave_time)

        waves_dict[wave_time][color].append({
            "route": route,
            "staging": staging,
            "dsp": dsp
        })

    print(f"  Rows processed: {row_num - data_start}  |  Skipped: {skipped}")
    print(f"  Waves found: {wave_order}")

    # Build final list sorted by wave time
    def wave_sort_key(t):
        m = re.match(r'(\d+):(\d+)\s*(AM|PM)', t, re.IGNORECASE)
        if not m:
            return 9999
        h, mn, ap = int(m.group(1)), int(m.group(2)), m.group(3).upper()
        if ap == "PM" and h != 12:
            h += 12
        if ap == "AM" and h == 12:
            h = 0
        return h * 60 + mn

    wave_order.sort(key=wave_sort_key)

    result = []
    for wt in wave_order:
        green = waves_dict[wt]["green"]
        red   = waves_dict[wt]["red"]
        result.append({
            "time":  wt,
            "total": len(green) + len(red),
            "green": green,
            "red":   red
        })

    return result

# ── JS serialiser ────────────────────────────────────────────────────────────

def route_to_js(r: dict) -> str:
    return '{' + f'route:"{r["route"]}",staging:"{r["staging"]}",dsp:"{r["dsp"]}"' + '}'

def waves_to_js(waves: list) -> str:
    lines = []
    lines.append("// Generated by import_data.py — " + datetime.now().strftime("%Y-%m-%d %H:%M"))
    lines.append("// DO NOT edit manually — re-run the script to update.")
    lines.append("")
    lines.append("const WAVES = [")

    for i, w in enumerate(waves):
        green_js = ",".join(route_to_js(r) for r in w["green"])
        red_js   = ",".join(route_to_js(r) for r in w["red"])
        comma = "," if i < len(waves) - 1 else ""
        lines.append(f'  {{')
        lines.append(f'    time: "{w["time"]}", total: {w["total"]},')
        lines.append(f'    green: [{green_js}],')
        lines.append(f'    red:   [{red_js}]')
        lines.append(f'  }}{comma}')

    lines.append("];")
    return "\n".join(lines)

# ── Entry point ──────────────────────────────────────────────────────────────

def main():
    excel_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_EXCEL
    sheet_name = sys.argv[2] if len(sys.argv) > 2 else None

    print("=" * 50)
    print("  DNX3 Wave Tracker — Excel → data.js")
    print("=" * 50)
    print(f"  Input:  {excel_path}")
    print(f"  Output: {OUTPUT_JS}")
    print()

    if not os.path.isfile(excel_path):
        print(f"ERROR: File not found: {excel_path}")
        print()
        print("Usage:")
        print("  python import_data.py                    # uses sequencing.xlsx")
        print("  python import_data.py myfile.xlsx")
        print("  python import_data.py myfile.xlsx Sheet2")
        sys.exit(1)

    print("Reading Excel...")
    try:
        waves = excel_to_waves(excel_path, sheet_name)
    except Exception as e:
        print(f"\nERROR reading Excel: {e}")
        sys.exit(1)

    if not waves:
        print("ERROR: No wave data found in the file.")
        sys.exit(1)

    # Summary
    total_routes = sum(w["total"] for w in waves)
    print()
    print("Summary:")
    for w in waves:
        print(f"  Wave {w['time']:10s}  —  {w['total']:3d} routes  "
              f"(🟢 {len(w['green'])} B/D  +  🔴 {len(w['red'])} A/C)")
    print(f"  {'TOTAL':14s}     {total_routes} routes across {len(waves)} waves")
    print()

    # Backup existing data.js
    if os.path.isfile(OUTPUT_JS):
        bak = OUTPUT_JS + ".bak"
        shutil.copy2(OUTPUT_JS, bak)
        print(f"  Backup saved: data.js.bak")

    # Write new data.js
    js_content = waves_to_js(waves)
    with open(OUTPUT_JS, "w", encoding="utf-8") as f:
        f.write(js_content)

    print(f"  ✅ data.js written successfully ({len(js_content):,} bytes)")
    print()
    print("Done! Refresh app.html in your browser to load the new data.")
    print("=" * 50)

if __name__ == "__main__":
    main()
