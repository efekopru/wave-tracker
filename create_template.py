"""
Creates sequencing_template.xlsx — a ready-to-fill Excel template
for the DNX3 Wave Tracker import script.

Run once:  python create_template.py
"""
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl not installed.  Run:  pip install openpyxl")
    raise

import os

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(SCRIPT_DIR, "sequencing_template.xlsx")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sequencing"

# ── Column widths ──
ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 10
ws.column_dimensions["E"].width = 16

# ── Header row ──
headers = ["Wave Time", "Route Code", "Staging", "DSP", "Staging Group"]
header_fill   = PatternFill("solid", fgColor="1E2D3D")
header_font   = Font(bold=True, color="FFFFFF", size=10)
header_align  = Alignment(horizontal="center", vertical="center")
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_idx, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font   = header_font
    cell.fill   = header_fill
    cell.alignment = header_align
    cell.border = border

ws.row_dimensions[1].height = 22

# ── Instruction row ──
instructions = [
    'e.g. "10:50 AM"',
    'e.g. "CA_A145"',
    'e.g. "STG-D1-21.6"',
    'e.g. "DERD"',
    'B, D = green  |  A, C = red  (optional — auto-detected if blank)'
]
inst_font = Font(italic=True, color="888888", size=9)
inst_fill = PatternFill("solid", fgColor="F9FAFB")
for col_idx, txt in enumerate(instructions, start=1):
    cell = ws.cell(row=2, column=col_idx, value=txt)
    cell.font   = inst_font
    cell.fill   = inst_fill
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = border

ws.row_dimensions[2].height = 18

# ── Sample data rows (first wave from current data) ──
sample = [
    ("10:50 AM", "CA_A145", "STG-D1-21.6",  "DERD", "D"),
    ("10:50 AM", "CA_A147", "STG-D1-21.3",  "DERD", "D"),
    ("10:50 AM", "CA_A148", "STG-D1-21.2",  "DERD", "D"),
    ("10:50 AM", "CA_A153", "STG-C1-28.11", "DERD", "C"),
    ("10:50 AM", "CA_A212", "STG-B1-30.1",  "LMDD", "B"),
    ("10:50 AM", "CA_A213", "STG-A1-28.1",  "LMDD", "A"),
    ("11:10 AM", "CA_A104", "STG-D1-21.6",  "ALGK", "D"),
    ("11:10 AM", "CA_A137", "STG-B1-30.10", "DERD", "B"),
    ("11:10 AM", "CA_A150", "STG-A1-28.11", "DERD", "A"),
]

green_fill  = PatternFill("solid", fgColor="D4EDDA")
red_fill    = PatternFill("solid", fgColor="FDE8E8")
green_font  = Font(size=10, color="1A7F3C")
red_font    = Font(size=10, color="B91C1C")
normal_font = Font(size=10)
normal_fill = PatternFill("solid", fgColor="FFFFFF")

for row_idx, (wave, route, staging, dsp, grp) in enumerate(sample, start=3):
    is_green = grp.upper() in ("B", "D")
    fill = green_fill if is_green else red_fill
    rfont = green_font if is_green else red_font
    vals = [wave, route, staging, dsp, grp]
    for col_idx, val in enumerate(vals, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.fill   = fill if col_idx in (2, 3, 4, 5) else normal_fill
        cell.font   = rfont if col_idx in (2, 5) else normal_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border
    ws.row_dimensions[row_idx].height = 16

# ── Notes sheet ──
notes = wb.create_sheet("Instructions")
notes.column_dimensions["A"].width = 70
note_lines = [
    ("DNX3 Wave Tracker — Sequencing Template Instructions", True),
    ("", False),
    ("COLUMNS", True),
    ("Wave Time    — Departure time of the wave, e.g. '10:50 AM' or '11:10 AM'", False),
    ("Route Code   — Route identifier, e.g. 'CA_A145'", False),
    ("Staging      — Full staging location code, e.g. 'STG-D1-21.6'", False),
    ("DSP          — DSP short code, e.g. 'DERD', 'LMDD', 'ALGK'", False),
    ("Staging Group — A, B, C or D (optional — auto-detected from Staging if left blank)", False),
    ("", False),
    ("STAGING GROUP RULES", True),
    ("  B or D  →  Green card  (Staging B & D section in the tracker)", False),
    ("  A or C  →  Red card    (Staging A & C section in the tracker)", False),
    ("", False),
    ("HOW TO USE", True),
    ("1. Fill in the 'Sequencing' sheet — one row per route", False),
    ("2. Delete the instruction row (row 2) and sample rows before importing", False),
    ("   OR leave them — the script skips rows with invalid data automatically", False),
    ("3. Save the file as sequencing.xlsx in the wave-tracker folder", False),
    ("4. Run:  python import_data.py", False),
    ("5. Refresh app.html in your browser", False),
    ("", False),
    ("CUSTOM FILE NAME", True),
    ("  python import_data.py  myfile.xlsx", False),
    ("  python import_data.py  myfile.xlsx  SheetName", False),
    ("", False),
    ("ACCEPTED COLUMN HEADER VARIATIONS", True),
    ("  Wave Time:     wave time / wave / departure / wave_time", False),
    ("  Route Code:    route code / route / route_code / tour / tour code", False),
    ("  Staging:       staging / staging location / stage / position", False),
    ("  DSP:           dsp / dsp code / provider / carrier", False),
    ("  Staging Group: staging group / group / color / zone  (all optional)", False),
]

hdr_font  = Font(bold=True, size=11, color="1E2D3D")
body_font = Font(size=10)
for row_idx, (text, bold) in enumerate(note_lines, start=1):
    cell = notes.cell(row=row_idx, column=1, value=text)
    cell.font = hdr_font if bold else body_font
    cell.alignment = Alignment(wrap_text=True)

wb.save(OUT)
print(f"✅ Template created: {OUT}")
print("   Fill in the 'Sequencing' sheet, then run: python import_data.py")
